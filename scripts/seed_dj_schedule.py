import argparse
import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, datetime, time, timedelta

from openpyxl import load_workbook


WEEKDAYS = {
    "MONDAY": 0,
    "TUESDAY": 1,
    "WEDNESDAY": 2,
    "THURSDAY": 3,
    "FRIDAY": 4,
    "SATURDAY": 5,
    "SUNDAY": 6,
}

ROLE_ORDER = ["Warm-up", "Driver", "Peak", "Closer"]


def clean(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def truthy_status(value):
    if isinstance(value, bool):
        return value
    text = clean(value).lower()
    return text in {"true", "yes", "y", "confirmed", "confirm", "done", "booked"}


def parse_month_from_sheet(sheet_name):
    parsed = datetime.strptime(sheet_name.title(), "%B %Y")
    return parsed.year, parsed.month


def infer_date(year, month, weekday_name, used_dates):
    target = WEEKDAYS.get(weekday_name)
    if target is None:
        return None

    candidate = date(year, month, 1)
    while candidate.month == month:
        if candidate.weekday() == target and candidate not in used_dates:
            return candidate
        candidate += timedelta(days=1)
    return None


def normalize_time_piece(piece):
    text = clean(piece).upper()
    text = text.replace(".", "").replace(" ", "")
    match = re.match(r"^(\d{1,2})(?::(\d{2}))?(AM|PM)$", text)
    if not match:
        return None

    hour = int(match.group(1))
    minute = int(match.group(2) or "0")
    suffix = match.group(3)
    if suffix == "AM":
        if hour == 12:
            hour = 0
    elif hour != 12:
        hour += 12
    return time(hour % 24, minute)


def parse_time_range(value):
    text = clean(value)
    if not text:
        return None, None
    text = text.replace("–", "-").replace("—", "-").replace("~", "-")
    parts = [part.strip() for part in text.split("-") if part.strip()]
    if len(parts) != 2:
        return None, None
    return normalize_time_piece(parts[0]), normalize_time_piece(parts[1])


def role_for_slot(index, total):
    if total <= 1:
        return "Peak"
    if index == 0:
        return "Warm-up"
    if index == total - 1:
        return "Closer"
    if total == 3:
        return "Peak"
    return "Driver" if index < total - 2 else "Peak"


def energy_for_role(role):
    return {"Warm-up": 2, "Driver": 3, "Peak": 4, "Closer": 4}.get(role, 3)


def event_status(slots):
    if slots and all(slot["confirmed"] for slot in slots):
        return "Confirmed"
    if slots:
        return "Unconfirmed"
    return "No Lineup"


def parse_workbook(path):
    wb = load_workbook(path, data_only=True)
    events = []

    for ws in wb.worksheets:
        try:
            year, month = parse_month_from_sheet(ws.title)
        except ValueError:
            continue

        rows = list(ws.iter_rows(values_only=True))
        used_dates = set()
        current_weekday = ""
        i = 0
        while i < len(rows):
            row = rows[i]
            b = clean(row[1] if len(row) > 1 else None)
            b_upper = str(b).upper()
            weekday = next((name for name in WEEKDAYS if b_upper.startswith(name)), None)
            if weekday:
                current_weekday = weekday
                i += 1
                continue

            headers = [str(clean(cell)).lower() for cell in row]
            if "date" in headers and "dj" in headers:
                date_idx = headers.index("date")
                time_idx = next((idx for idx, header in enumerate(headers) if header == "time"), None)
                dj_idx = headers.index("dj")
                status_idx = next((idx for idx, header in enumerate(headers) if header == "status"), None)
                remarks_idx = next((idx for idx, header in enumerate(headers) if header == "remarks"), None)
                genre_idx = next((idx for idx, header in enumerate(headers) if header == "genre"), None)
                stage_idx = next((idx for idx, header in enumerate(headers) if header == "stage"), None)

                event_name = ""
                for back in range(i - 1, max(-1, i - 5), -1):
                    candidate = clean(rows[back][1] if len(rows[back]) > 1 else None)
                    if candidate and not str(candidate).upper().startswith(("WEEK ", *WEEKDAYS.keys())):
                        event_name = candidate
                        break

                block = []
                j = i + 1
                while j < len(rows):
                    next_row = rows[j]
                    next_b = clean(next_row[1] if len(next_row) > 1 else None)
                    next_headers = [str(clean(cell)).lower() for cell in next_row]
                    if "date" in next_headers and "dj" in next_headers:
                        break
                    if str(next_b).upper().startswith(("WEEK ", *WEEKDAYS.keys())):
                        break

                    dj = clean(next_row[dj_idx] if dj_idx < len(next_row) else None)
                    slot_time = clean(next_row[time_idx] if time_idx is not None and time_idx < len(next_row) else None)
                    row_date = next_row[date_idx] if date_idx < len(next_row) else None
                    genre = clean(next_row[genre_idx] if genre_idx is not None and genre_idx < len(next_row) else None)
                    stage = clean(next_row[stage_idx] if stage_idx is not None and stage_idx < len(next_row) else None)
                    remarks = clean(next_row[remarks_idx] if remarks_idx is not None and remarks_idx < len(next_row) else None)
                    confirmed = truthy_status(next_row[status_idx] if status_idx is not None and status_idx < len(next_row) else None)

                    if dj or slot_time or isinstance(row_date, (datetime, date)) or genre or stage or remarks:
                        block.append(
                            {
                                "date": row_date,
                                "time": slot_time,
                                "dj": dj,
                                "confirmed": confirmed,
                                "remarks": remarks,
                                "genre": genre,
                                "stage": stage,
                            }
                        )
                    j += 1

                event_date = next((entry["date"] for entry in block if isinstance(entry["date"], (datetime, date))), None)
                if isinstance(event_date, datetime):
                    event_date = event_date.date()
                if event_date is None:
                    event_date = infer_date(year, month, current_weekday, used_dates)

                if event_name and event_date:
                    used_dates.add(event_date)
                    genre = next((entry["genre"] for entry in block if entry["genre"]), "")
                    stage = next((entry["stage"] for entry in block if entry["stage"]), "") or "Centre Stage"
                    notes = "; ".join(dict.fromkeys(entry["remarks"] for entry in block if entry["remarks"]))
                    slots = []
                    timed_entries = []

                    for entry in block:
                        if not entry["dj"]:
                            continue
                        start, end = parse_time_range(entry["time"])
                        timed_entries.append({**entry, "start": start, "end": end})

                    total = len(timed_entries)
                    for idx, entry in enumerate(timed_entries):
                        role = role_for_slot(idx, total)
                        slots.append(
                            {
                                "slot_order": idx + 1,
                                "start_time": entry["start"].strftime("%H:%M:%S") if entry["start"] else "22:00:00",
                                "end_time": entry["end"].strftime("%H:%M:%S") if entry["end"] else "03:30:00",
                                "role": role,
                                "expected_energy": energy_for_role(role),
                                "dj_name": entry["dj"],
                                "assignment_status": "Confirmed" if entry["confirmed"] else "Pending",
                                "assignment_notes": entry["remarks"] or None,
                                "confirmed": entry["confirmed"],
                            }
                        )

                    events.append(
                        {
                            "event_name": event_name,
                            "event_date": event_date.isoformat(),
                            "day_name": current_weekday.title() if current_weekday else event_date.strftime("%A"),
                            "event_type": event_name,
                            "genre_profile": genre or None,
                            "stage": stage,
                            "status": event_status(slots),
                            "notes": notes or None,
                            "slots": slots,
                            "source_sheet": ws.title,
                        }
                    )

                i = j
                continue
            i += 1

    return events


class SupabaseRest:
    def __init__(self, url, key):
        self.base = url.rstrip("/") + "/rest/v1"
        self.key = key

    def request(self, method, path, body=None, query=None, prefer=None):
        qs = ("?" + urllib.parse.urlencode(query, doseq=True)) if query else ""
        req = urllib.request.Request(self.base + path + qs, method=method)
        req.add_header("apikey", self.key)
        req.add_header("Authorization", f"Bearer {self.key}")
        req.add_header("Content-Type", "application/json")
        if prefer:
            req.add_header("Prefer", prefer)
        data = json.dumps(body).encode("utf-8") if body is not None else None
        try:
            with urllib.request.urlopen(req, data=data, timeout=30) as response:
                text = response.read().decode("utf-8")
                return json.loads(text) if text else None
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8")
            raise RuntimeError(f"{method} {path} failed: {exc.code} {detail}") from exc

    def find_event(self, event):
        query = {
            "select": "id",
            "event_date": f"eq.{event['event_date']}",
            "event_name": f"eq.{event['event_name']}",
            "limit": "1",
        }
        rows = self.request("GET", "/events", query=query) or []
        return rows[0]["id"] if rows else None

    def upsert_event(self, event):
        existing_id = self.find_event(event)
        payload = {key: event[key] for key in ["event_name", "event_date", "day_name", "event_type", "genre_profile", "stage", "status", "notes"]}
        if existing_id:
            rows = self.request("PATCH", "/events", payload, query={"id": f"eq.{existing_id}", "select": "id"}, prefer="return=representation")
            return rows[0]["id"], False
        rows = self.request("POST", "/events", payload, query={"select": "id"}, prefer="return=representation")
        return rows[0]["id"], True

    def get_or_create_dj(self, name):
        query = {"select": "id", "name": f"eq.{name}", "limit": "1"}
        rows = self.request("GET", "/djs", query=query) or []
        if rows:
            return rows[0]["id"], False
        try:
            rows = self.request("POST", "/djs", {"name": name}, query={"select": "id"}, prefer="return=representation")
            return rows[0]["id"], True
        except RuntimeError as error:
            if "row-level security policy" in str(error):
                return None, False
            raise

    def create_slot_assignment(self, event_id, slot):
        slot_payload = {key: slot[key] for key in ["slot_order", "start_time", "end_time", "role", "expected_energy"]}
        slot_payload["event_id"] = event_id
        existing_slots = self.request(
            "GET",
            "/event_slots",
            query={"select": "id", "event_id": f"eq.{event_id}", "slot_order": f"eq.{slot['slot_order']}", "limit": "1"},
        ) or []
        if existing_slots:
            slot_id = existing_slots[0]["id"]
            self.request("PATCH", "/event_slots", slot_payload, query={"id": f"eq.{slot_id}"})
        else:
            rows = self.request("POST", "/event_slots", slot_payload, query={"select": "id"}, prefer="return=representation")
            slot_id = rows[0]["id"]

        existing_assignments = self.request(
            "GET",
            "/event_assignments",
            query={"select": "id", "event_slot_id": f"eq.{slot_id}", "limit": "1"},
        ) or []
        dj_id, created_dj = self.get_or_create_dj(slot["dj_name"])
        assignment_payload = {
            "event_slot_id": slot_id,
            "dj_id": dj_id,
            "assignment_status": slot["assignment_status"],
            "notes": slot["assignment_notes"] or (None if dj_id else slot["dj_name"]),
        }
        if existing_assignments:
            self.request("PATCH", "/event_assignments", assignment_payload, query={"id": f"eq.{existing_assignments[0]['id']}"})
        else:
            self.request("POST", "/event_assignments", assignment_payload)
        return created_dj


def read_env(path=".env.local"):
    env = {}
    with open(path, "r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            env[key] = value
    return env


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    events = parse_workbook(args.file)
    print(json.dumps({"events": len(events), "slots": sum(len(event["slots"]) for event in events), "djs": len({slot["dj_name"] for event in events for slot in event["slots"]})}, indent=2))
    print(json.dumps(events[:5], indent=2))

    if not args.apply:
        return

    env = read_env()
    supabase = SupabaseRest(env["VITE_SUPABASE_URL"], env["VITE_SUPABASE_ANON_KEY"])
    stats = {"created_events": 0, "updated_events": 0, "created_djs": 0, "slots": 0}
    for event in events:
        event_id, created_event = supabase.upsert_event(event)
        if created_event:
            stats["created_events"] += 1
        else:
            stats["updated_events"] += 1
        for slot in event["slots"]:
            if supabase.create_slot_assignment(event_id, slot):
                stats["created_djs"] += 1
            stats["slots"] += 1
    print(json.dumps(stats, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        print(error, file=sys.stderr)
        sys.exit(1)
