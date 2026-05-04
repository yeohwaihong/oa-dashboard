#!/usr/bin/env python3
import argparse
import calendar
import json
import os
import re
import sys
import urllib.error
import urllib.request
import urllib.parse
from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


COLUMNS = [
    "date",
    "event_name",
    "pax",
    "table_bookings",
    "door_sales",
    "cover_charge",
    "pos_total",
    "ticketmelon_total",
    "weekly_target",
    "week_number",
    "month_year",
    "utilities",
    "man_power",
    "local_dj",
    "ambassador_commission",
    "intl_artist_cost",
    "puspal",
    "hotel",
    "rider",
    "bar_split",
    "misc",
    "online_ticket",
    "walkin_ticket",
    "sponsorship",
]


MONTH_NAMES = {name.upper(): index for index, name in enumerate(calendar.month_name) if name}
MONTH_NAMES.update({name.upper(): index for index, name in enumerate(calendar.month_abbr) if name})


def clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value).strip()
    return value


def num(value):
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    try:
        return float(text)
    except ValueError:
        return 0.0


def iso(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return ""


def sheet_month_year(sheet_name):
    match = re.search(r"SALES\s+(\d{4})\s*-\s*([A-Z]+)", sheet_name.upper())
    if not match:
        return ""
    year = match.group(1)
    month = match.group(2)
    month_num = MONTH_NAMES.get(month)
    month_label = calendar.month_name[month_num].upper() if month_num else month
    return f"{month_label} {year}"


def default_costs_for_date(date_iso):
    weekday = datetime.strptime(date_iso, "%Y-%m-%d").weekday()
    is_fri_sat = weekday in (4, 5)
    is_sat = weekday == 5
    return {
        "utilities": 1500,
        "man_power": 13100 if is_fri_sat else 9500,
        "local_dj": 2300 if is_sat else 1500 if weekday in (2, 4) else 2000,
    }


def normalize_record(raw):
    costs = default_costs_for_date(raw["date"])
    return {
        "date": raw["date"],
        "event_name": raw["event_name"],
        "pax": raw["pax"],
        "table_bookings": round(num(raw.get("table_bookings")), 2),
        "door_sales": round(num(raw.get("door_sales")), 2),
        "cover_charge": round(num(raw.get("cover_charge")), 2),
        "pos_total": round(num(raw.get("pos_total")), 2),
        "ticketmelon_total": round(num(raw.get("ticketmelon_total")), 2),
        "weekly_target": round(num(raw.get("weekly_target")), 2),
        "week_number": int(raw["week_number"]),
        "month_year": raw["month_year"],
        **costs,
        "ambassador_commission": 0,
        "intl_artist_cost": 0,
        "puspal": 0,
        "hotel": 0,
        "rider": 0,
        "bar_split": 0,
        "misc": 0,
        "online_ticket": 0,
        "walkin_ticket": 0,
        "sponsorship": 0,
    }


def merge_records(records):
    by_date = OrderedDict()
    for record in records:
        key = record["date"]
        if key not in by_date:
            by_date[key] = normalize_record(record)
            continue

        existing = by_date[key]
        event_name = clean(record.get("event_name"))
        pax = clean(record.get("pax"))
        if event_name and event_name not in existing["event_name"]:
            existing["event_name"] = f"{existing['event_name']} + {event_name}"
        if pax:
            existing["pax"] = f"{existing['pax']} / {pax}" if existing["pax"] else str(pax)
        for field in ["table_bookings", "door_sales", "cover_charge", "pos_total", "ticketmelon_total"]:
            existing[field] = round(existing[field] + num(record.get(field)), 2)
        if not existing["weekly_target"] and num(record.get("weekly_target")):
            existing["weekly_target"] = round(num(record.get("weekly_target")), 2)
    return list(by_date.values())


def parse_workbook(path):
    workbook = load_workbook(path, data_only=True, read_only=False)
    records = []

    for worksheet in workbook.worksheets:
        if not worksheet.title.upper().startswith("SALES 20"):
            continue
        month_year = sheet_month_year(worksheet.title)
        week_number = 0
        in_table = False
        last_date = ""

        for row in worksheet.iter_rows(values_only=True):
            cells = list(row[:10])
            first = clean(cells[0] if len(cells) > 0 else "")
            second = clean(cells[1] if len(cells) > 1 else "")

            if str(first).upper() in {"DATE", "WEEK COMMENCING"}:
                week_number += 1
                in_table = True
                last_date = ""
                continue
            if str(first).upper().startswith("GRAND TOTAL"):
                in_table = False
                last_date = ""
                continue
            if not in_table:
                continue

            current_date = iso(cells[0] if len(cells) > 0 else None)
            if current_date:
                last_date = current_date
                event_name = second
                offset = 0
            elif last_date and second:
                current_date = last_date
                event_name = second
                offset = 0
            else:
                continue

            if not event_name or "GRAND TOTAL" in str(event_name).upper():
                continue

            records.append(
                {
                    "date": current_date,
                    "event_name": str(event_name),
                    "pax": str(clean(cells[2 + offset])) if len(cells) > 2 + offset and clean(cells[2 + offset]) != "" else "0",
                    "table_bookings": cells[3 + offset] if len(cells) > 3 + offset else 0,
                    "door_sales": cells[4 + offset] if len(cells) > 4 + offset else 0,
                    "cover_charge": cells[5 + offset] if len(cells) > 5 + offset else 0,
                    "pos_total": cells[6 + offset] if len(cells) > 6 + offset else 0,
                    "ticketmelon_total": cells[7 + offset] if len(cells) > 7 + offset else 0,
                    "weekly_target": cells[9 + offset] if len(cells) > 9 + offset else 0,
                    "week_number": week_number,
                    "month_year": month_year,
                }
            )

    return merge_records(records)


def sql_value(value):
    if value is None:
        return "null"
    if isinstance(value, (int, float)):
        return str(value)
    return "'" + str(value).replace("'", "''") + "'"


def write_sql(records, path):
    months = sorted({record["month_year"] for record in records if record["month_year"]})
    values = []
    for record in records:
        values.append("  (" + ", ".join(sql_value(record[column]) for column in COLUMNS) + ")")

    update_columns = [column for column in COLUMNS if column != "date"]
    sql = [
        "-- Seed: Weekly Sales from OA WIP SHEET_UPDATED 16 JAN.xlsx",
        "-- Generated by scripts/seed_weekly_sales_from_workbook.py. Safe to re-run.",
        "",
        f"delete from public.weekly_sales where month_year in ({', '.join(sql_value(month) for month in months)});",
        "",
        "insert into public.weekly_sales",
        "  (" + ", ".join(COLUMNS) + ")",
        "values",
        ",\n".join(values),
        "on conflict (date) do update set",
        ",\n".join(f"  {column} = excluded.{column}" for column in update_columns) + ";",
        "",
    ]
    path.write_text("\n".join(sql), encoding="utf-8")


def read_env(paths):
    env = dict(os.environ)
    for path in paths:
        if not path.exists():
            continue
        for line in path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            env.setdefault(key.strip(), value.strip().strip('"').strip("'"))
    return env


def apply_to_supabase(records, env):
    url = env.get("SUPABASE_URL") or env.get("VITE_SUPABASE_URL")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        raise RuntimeError("Missing SUPABASE_URL/VITE_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY.")

    months = sorted({record["month_year"] for record in records if record["month_year"]})
    quoted_months = ",".join(f'"{month}"' for month in months)
    delete_endpoint = url.rstrip("/") + "/rest/v1/weekly_sales?month_year=in.(" + urllib.parse.quote(quoted_months, safe='(),"') + ")"
    delete_request = urllib.request.Request(
        delete_endpoint,
        method="DELETE",
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Prefer": "return=minimal",
        },
    )
    with urllib.request.urlopen(delete_request, timeout=60) as response:
        response.read()

    endpoint = url.rstrip("/") + "/rest/v1/weekly_sales?on_conflict=date"
    payload = json.dumps(records).encode("utf-8")
    request = urllib.request.Request(
        endpoint,
        data=payload,
        method="POST",
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates,return=minimal",
        },
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        response.read()


def main():
    parser = argparse.ArgumentParser(description="Extract OA WIP workbook sales data into weekly_sales seed rows.")
    parser.add_argument("workbook", help="Path to OA WIP workbook")
    parser.add_argument("--sql", default="supabase/seed_weekly_sales_from_oa_wip.sql", help="Output SQL path")
    parser.add_argument("--apply", action="store_true", help="Upsert rows directly to Supabase using service role env")
    args = parser.parse_args()

    records = parse_workbook(args.workbook)
    sql_path = Path(args.sql)
    write_sql(records, sql_path)
    print(json.dumps({"rows": len(records), "sql": str(sql_path), "first": records[0]["date"] if records else None, "last": records[-1]["date"] if records else None}, indent=2))

    if args.apply:
        env = read_env([Path(".env.local"), Path(".env.development.local"), Path(".env.production.local"), Path(".env")])
        try:
            apply_to_supabase(records, env)
            print(json.dumps({"applied": True, "rows": len(records)}, indent=2))
        except urllib.error.HTTPError as error:
            body = error.read().decode("utf-8", errors="replace")
            print(body, file=sys.stderr)
            raise


if __name__ == "__main__":
    main()
